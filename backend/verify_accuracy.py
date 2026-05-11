"""校验脚本：对比 Excel 原始数据 vs SQLite 库中数据"""
import sqlite3
from openpyxl import load_workbook
from datetime import datetime, timedelta

def serial_to_date(v):
    """Excel 序列号 → 日期字符串"""
    if isinstance(v, (int, float)) and 40000 < v < 60000:
        return (datetime(1899, 12, 30) + timedelta(days=int(v))).strftime('%Y-%m-%d')
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    return str(v)[:10] if v else None

# ===== 1. 专业项目校验 =====
print("=" * 60)
print("1. 专业项目对比")
print("=" * 60)
wb = load_workbook('../四项规划/后勤部-专业.xlsx')
ws = wb[wb.sheetnames[0]]

excel_proj = {}
for i in range(1, ws.max_row + 1):
    seq = ws.cell(row=i, column=1).value
    if isinstance(seq, int) and seq >= 1:
        dept = ws.cell(row=i, column=2).value
        if dept is None:
            continue  # 跳过空部门行（说明该行只有序号没有数据）
        n = ws.max_column
        # 从右往左取：duration, end_date, start_date, person, deliverable
        end_d = serial_to_date(ws.cell(row=i, column=n-1).value)  # col-2
        start_d = serial_to_date(ws.cell(row=i, column=n-2).value)  # col-3
        person = ws.cell(row=i, column=n-3).value
        deliverable = ws.cell(row=i, column=n-4).value
        duration_v = ws.cell(row=i, column=n).value
        duration = str(duration_v) if duration_v else None
        
        # 计算阶段数
        pc = 0
        for j in range(5, 12):
            v = ws.cell(row=i, column=j).value
            if v and '阶段' in str(v)[:10]:
                pc += 1
            else:
                break
        
        excel_proj[seq] = {
            'dept': dept, 'name': ws.cell(row=i, column=3).value,
            'person': person, 'start': start_d, 'end': end_d,
            'duration': duration, 'phases': pc
        }

conn = sqlite3.connect('data/houqin.db')
cur = conn.cursor()
cur.execute("SELECT id, department, name, person, start_date, end_date, duration, phase_count FROM professional_projects ORDER BY id")
db_proj = {r[0]: {'dept': r[1], 'name': r[2], 'person': r[3], 'start': r[4], 'end': r[5], 'duration': r[6], 'phases': r[7]} for r in cur.fetchall()}

issues_pro = []
for seq in sorted(set(list(excel_proj.keys()) + list(db_proj.keys()))):
    e = excel_proj.get(seq)
    d = db_proj.get(seq)
    if e and d:
        person_ok = str(e['person']) == str(d['person'])
        date_ok = str(e['start'])[:10] == str(d['start'])[:10] and str(e['end'])[:10] == str(d['end'])[:10]
        phase_ok = e['phases'] == d['phases']
        status = '✓' if (person_ok and date_ok and phase_ok) else '✗'
        if status == '✗':
            print(f"  [✗] id={seq}: {e['name']}")
            if not person_ok: print(f"       person: Excel={e['person']} vs DB={d['person']}")
            if not date_ok: print(f"       date: Excel={e['start']}~{e['end']} vs DB={d['start']}~{d['end']}")
            if not phase_ok: print(f"       phases: Excel={e['phases']} vs DB={d['phases']}")
    elif e and not d:
        print(f"  [✗] id={seq}: 只存在于Excel，DB缺失! name={e['name']}")
    elif d and not e:
        print(f"  [✗] id={seq}: 只存在于DB，Excel缺失! name={d['name']}")

print(f"  结果: Excel={len(excel_proj)}项, DB={len(db_proj)}项")

# ===== 2. IT项目校验 =====
print()
print("=" * 60)
print("2. IT项目对比")
print("=" * 60)
wb2 = load_workbook('../四项规划/后勤部-信息化.xlsx')
ws2 = wb2[wb2.sheetnames[0]]

excel_it = {}
for i in range(1, ws2.max_row + 1):
    seq = ws2.cell(row=i, column=1).value
    cat = ws2.cell(row=i, column=2).value
    if isinstance(seq, int) and seq >= 1 and cat:
        n = ws2.max_column
        name = ws2.cell(row=i, column=3).value
        owner = ws2.cell(row=i, column=n-5).value  # col-6
        pc = 0
        for j in range(5, 12):
            v = ws2.cell(row=i, column=j).value
            if v and '阶段' in str(v)[:10]:
                pc += 1
            else:
                break
        excel_it[seq] = {'cat': cat, 'name': name, 'owner': owner, 'phases': pc}

cur.execute("SELECT id, category, name, owner, phase_count FROM it_projects ORDER BY id")
db_it = {r[0]: {'cat': r[1], 'name': r[2], 'owner': r[3], 'phases': r[4]} for r in cur.fetchall()}

for seq in sorted(set(list(excel_it.keys()) + list(db_it.keys()))):
    e = excel_it.get(seq)
    d = db_it.get(seq)
    if e and d:
        name_ok = str(e['name'])[:30] == str(d['name'])[:30]
        phase_ok = e['phases'] == d['phases']
        status = '✓' if (name_ok and phase_ok) else '✗'
        if status == '✗':
            print(f"  [✗] id={seq}: IT Excel={e['name'][:30]}, DB={d['name'][:30]} phases Excel={e['phases']} DB={d['phases']}")
    elif e and not d:
        print(f"  [✗] id={seq}: 只存在于Excel，DB缺失! name={e['name'][:30]}")
    elif d and not e:
        print(f"  [✗] id={seq}: 只存在于DB，Excel缺失! name={d['name'][:30]}")

print(f"  结果: Excel={len(excel_it)}项, DB={len(db_it)}项")

# ===== 3. 人员名册校验 =====
print()
print("=" * 60)
print("3. 人员名册对比")
print("=" * 60)
wb3 = load_workbook('../四项规划/后勤部-人力.xlsx')
ws3 = wb3['人员调整计划']
excel_emp = 0
for i in range(3, ws3.max_row + 1):
    n = ws3.cell(row=i, column=2).value
    if n:
        excel_emp += 1

cur.execute("SELECT COUNT(*) FROM employees")
db_emp = cur.fetchone()[0]
cur.execute("SELECT COUNT(*) FROM employee_monthly_status")
db_monthly = cur.fetchone()[0]

print(f"  人员: Excel={excel_emp}人, DB={db_emp}人 | 月度状态: DB={db_monthly}条")

# 抽查前3人
for i in [3, 4, 5]:
    en = ws3.cell(row=i, column=2).value
    ep = ws3.cell(row=i, column=6).value
    eedu = ws3.cell(row=i, column=25).value
    cur.execute("SELECT name, post, education FROM employees WHERE name=?", (en,))
    r = cur.fetchone()
    if r:
        ok = (r[0] == en and str(r[1]) == str(ep) and str(r[2]) == str(eedu))
        print(f"  {'✓' if ok else '✗'} {en}: post Excel={ep} DB={r[1]}, edu Excel={eedu} DB={r[2]}")
    else:
        print(f"  [✗] {en}: 在DB中未找到!")

# ===== 4. 财务校验 =====
print()
print("=" * 60)
print("4. 财务数据对比")
print("=" * 60)
wb4 = load_workbook('../四项规划/后勤部-财务.xlsx')
ws_fi = wb4['宏观视角与目标']
for i in [4, 5, 6, 7, 8]:
    indicator = ws_fi.cell(row=i, column=1).value
    value = ws_fi.cell(row=i, column=3).value
    unit = ws_fi.cell(row=i, column=4).value
    if indicator and value:
        cur.execute("SELECT value, unit FROM financial_indicators WHERE indicator_name=?", (indicator,))
        r = cur.fetchone()
        if r:
            ok = abs(float(r[0]) - float(value)) < 0.1
            print(f"  {'✓' if ok else '✗'} {indicator}: Excel={value} DB={r[0]}")

cur.execute("SELECT COUNT(*) FROM financial_timeline")
print(f"  执行时间线: DB={cur.fetchone()[0]}条")
cur.execute("SELECT COUNT(*) FROM financial_reduction")
print(f"  降费方案: DB={cur.fetchone()[0]}条")
cur.execute("SELECT COUNT(*) FROM financial_budget")
print(f"  预算表: DB={cur.fetchone()[0]}条")

conn.close()
print()
print("=" * 60)
print("校验完成!")
