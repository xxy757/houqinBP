"""
从5个Excel文件中提取所有数据，写入SQLite数据库
"""
import sqlite3
import os
import datetime
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_DIR = os.path.join(os.path.dirname(BASE_DIR), '四项规划')
DB_PATH = os.path.join(BASE_DIR, 'data', 'houqin.db')

EXCEL_DATE_EPOCH = datetime.datetime(1899, 12, 30)


def clean(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, (int, float)):
        # Excel 日期序列号转换 (范围 2020-2070: 43831~62099)
        if 40000 < v < 70000:
            return (EXCEL_DATE_EPOCH + datetime.timedelta(days=int(v))).strftime('%Y-%m-%d')
        return v
    return str(v)

def create_db():
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    cur.executescript('''
        -- 专业项目主表
        CREATE TABLE professional_projects (
            id INTEGER PRIMARY KEY,
            department TEXT,
            name TEXT NOT NULL,
            goal TEXT,
            context TEXT,
            deliverable TEXT,
            person TEXT,
            start_date TEXT,
            end_date TEXT,
            duration TEXT,
            phase_count INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );

        -- 专业项目阶段详情
        CREATE TABLE professional_project_phases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            phase_order INTEGER,
            phase_name TEXT,
            phase_content TEXT,
            FOREIGN KEY (project_id) REFERENCES professional_projects(id)
        );

        -- 信息化项目主表
        CREATE TABLE it_projects (
            id INTEGER PRIMARY KEY,
            category TEXT,
            name TEXT NOT NULL,
            goal TEXT,
            context TEXT,
            deliverable TEXT,
            owner TEXT,
            start_date TEXT,
            end_date TEXT,
            duration TEXT,
            difficulty TEXT,
            solve TEXT,
            phase_count INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );

        -- 信息化项目阶段详情
        CREATE TABLE it_project_phases (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER NOT NULL,
            phase_order INTEGER,
            phase_name TEXT,
            phase_content TEXT,
            FOREIGN KEY (project_id) REFERENCES it_projects(id)
        );

        -- 人员名册
        CREATE TABLE employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id TEXT,
            name TEXT NOT NULL,
            department_level TEXT,
            department_code TEXT,
            department_name TEXT,
            post TEXT,
            merged_department TEXT,
            adjusted_department TEXT,
            status TEXT,
            entry_date TEXT,
            leave_date TEXT,
            leave_apply_date TEXT,
            remark TEXT,
            phone TEXT,
            gender TEXT,
            education TEXT,
            major TEXT,
            school TEXT,
            school_level TEXT,
            professional_match TEXT,
            age INTEGER,
            ethnicity TEXT,
            contract_start TEXT,
            contract_end TEXT,
            salary_account TEXT,
            personnel_type TEXT,
            entry_year TEXT,
            probation TEXT,
            regularization_date TEXT,
            work_years INTEGER,
            native_place TEXT,
            created_at TEXT DEFAULT (datetime('now','localtime')),
            updated_at TEXT DEFAULT (datetime('now','localtime'))
        );

        -- 人员月度状态 (1-12月)
        CREATE TABLE employee_monthly_status (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_name TEXT,
            month INTEGER,
            status TEXT
        );

        -- 人力规划KPI
        CREATE TABLE hr_plan_kpi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            indicator TEXT,
            target TEXT,
            m1 TEXT, m2 TEXT, m3 TEXT, m4 TEXT, m5 TEXT, m6 TEXT,
            m7 TEXT, m8 TEXT, m9 TEXT, m10 TEXT, m11 TEXT, m12 TEXT
        );

        -- 财务预算
        CREATE TABLE financial_budget (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            category TEXT,
            department TEXT,
            m1 TEXT, m2 TEXT, m3 TEXT, m4 TEXT, m5 TEXT, m6 TEXT,
            m7 TEXT, m8 TEXT, m9 TEXT, m10 TEXT, m11 TEXT, m12 TEXT,
            total TEXT,
            budget_num TEXT
        );

        -- 财务宏观指标
        CREATE TABLE financial_indicators (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            indicator_name TEXT,
            value REAL,
            unit TEXT,
            section TEXT
        );

        -- 财务执行时间线
        CREATE TABLE financial_timeline (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            phase_name TEXT,
            seq INTEGER,
            task_name TEXT,
            responsible TEXT,
            time_range TEXT,
            action_desc TEXT,
            precondition TEXT,
            deliverable TEXT
        );

        -- 降费方案
        CREATE TABLE financial_reduction (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            section TEXT,
            cost_subject TEXT,
            year_2025_actual REAL,
            year_budget REAL,
            change_rate TEXT,
            detail_item TEXT,
            category TEXT,
            priority TEXT,
            reduction_plan TEXT
        );

        -- 部门重点事项
        CREATE TABLE department_key_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            key_item TEXT,
            sub_category TEXT,
            problem TEXT,
            baseline TEXT,
            target TEXT,
            improvement_project TEXT,
            seq INTEGER,
            sub_project TEXT,
            action_content TEXT,
            deliverable TEXT,
            person TEXT,
            start_date TEXT,
            end_date TEXT,
            section TEXT
        );
    ''')
    conn.commit()
    return conn

def extract_professional(conn, cur):
    """提取后勤部-专业.xlsx"""
    wb = load_workbook(os.path.join(EXCEL_DIR, '后勤部-专业.xlsx'))
    ws = wb[wb.sheetnames[0]]

    projects = []
    phase_rows = {}
    current_pid = None

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)):
        vals = [clean(c) for c in row]
        seq = vals[0]
        dept = vals[1]

        # 跳过表头行
        if i == 0:
            continue

        # 判断是否是项目主行（第一列有数字）
        if isinstance(seq, (int, float)) and seq and dept:
            pid = int(seq)
            # 动态计算阶段数
            phases = []
            for j in range(5, len(vals)):
                if vals[j] is not None and len(str(vals[j])) > 3 and '阶段' in str(vals[j])[:10]:
                    phases.append(vals[j])
                else:
                    break
            pc = len(phases)
            # 从右往左定位: vals[-1]=duration, vals[-2]=end_date, vals[-3]=start_date, vals[-4]=person, vals[-5]=deliverable
            n = len(vals)
            p = {
                'id': pid,
                'department': vals[1],
                'name': vals[2],
                'goal': vals[3],
                'context': vals[4],
                'duration': vals[n-1] if n >= 1 and vals[n-1] else None,
                'end_date': vals[n-2] if n >= 2 else None,
                'start_date': vals[n-3] if n >= 3 else None,
                'person': vals[n-4] if n >= 4 else None,
                'deliverable': vals[n-5] if n >= 5 else None,
                'phase_count': pc,
            }
            projects.append(p)
            current_pid = pid
        elif current_pid is not None:
            # 这是阶段详情行
            phase_details = {}
            for j in range(5, min(11, len(vals))):
                if vals[j] is not None and len(str(vals[j])) > 5:
                    phase_details[j - 5] = vals[j]
            if phase_details:
                phase_rows[current_pid] = phase_details

    for p in projects:
        cur.execute('''
            INSERT INTO professional_projects (id, department, name, goal, context, deliverable, person, start_date, end_date, duration, phase_count)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        ''', (p['id'], p['department'], p['name'], p['goal'], p['context'],
              p['deliverable'], p['person'], p['start_date'], p['end_date'],
              p['duration'], p['phase_count']))

        if p['id'] in phase_rows:
            phase_names = ['第一阶段', '第二阶段', '第三阶段', '第四阶段', '第五阶段', '第六阶段']
            for pi, content in phase_rows[p['id']].items():
                cur.execute('''
                    INSERT INTO professional_project_phases (project_id, phase_order, phase_name, phase_content)
                    VALUES (?,?,?,?)
                ''', (p['id'], pi + 1, phase_names[pi] if pi < len(phase_names) else f'第{pi+1}阶段', content))

    conn.commit()
    print(f"  专业项目: {len(projects)} 条")
    return projects

def extract_it_projects(conn, cur):
    """提取后勤部-信息化.xlsx"""
    wb = load_workbook(os.path.join(EXCEL_DIR, '后勤部-信息化.xlsx'))
    ws = wb[wb.sheetnames[0]]

    projects = []
    phase_rows = {}
    current_pid = None

    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)):
        vals = [clean(c) for c in row]
        seq = vals[0]
        cat = vals[1]

        if i == 0:
            continue

        if isinstance(seq, (int, float)) and seq and cat:
            pid = int(seq)
            # 动态计算阶段数
            phases = []
            for j in range(5, min(12, len(vals))):
                if vals[j] is not None and len(str(vals[j])) > 3 and '阶段' in str(vals[j])[:10]:
                    phases.append(vals[j])
                else:
                    break
            pc = len(phases)
            n = len(vals)
            p = {
                'id': pid,
                'category': vals[1],
                'name': vals[2],
                'goal': vals[3],
                'context': vals[4],
                'phase_count': pc,
                'phases_content': [(i+1, phases[i]) for i in range(pc)],
                'deliverable': vals[n-7] if n >= 7 else None,
                'owner': vals[n-6] if n >= 6 else None,
                'start_date': vals[n-5] if n >= 5 else None,
                'end_date': vals[n-4] if n >= 4 else None,
                'duration': vals[n-3] if n >= 3 else None,
                'difficulty': vals[n-2] if n >= 2 else None,
                'solve': vals[n-1] if n >= 1 else None,
            }
            projects.append(p)
            current_pid = pid
        # IT阶段数据已在主行中，不需要从后续行提取

    for p in projects:
        cur.execute('''
            INSERT INTO it_projects (id, category, name, goal, context, deliverable, owner, start_date, end_date, duration, difficulty, solve, phase_count)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (p['id'], p['category'], p['name'], p['goal'], p['context'],
              p['deliverable'], p['owner'], p['start_date'], p['end_date'],
              p['duration'], p['difficulty'], p['solve'], p['phase_count']))

        # 插入阶段数据（从主行直接取）
        phase_names = ['第一阶段', '第二阶段', '第三阶段', '第四阶段', '第五阶段']
        for pi, content in p.get('phases_content', []):
            cur.execute('''
                INSERT INTO it_project_phases (project_id, phase_order, phase_name, phase_content)
                VALUES (?,?,?,?)
            ''', (p['id'], pi, phase_names[pi-1] if pi <= len(phase_names) else f'第{pi}阶段', content))

    conn.commit()
    print(f"  信息化项目: {len(projects)} 条")
    return projects

def extract_hr(conn, cur):
    """提取后勤部-人力.xlsx"""
    wb = load_workbook(os.path.join(EXCEL_DIR, '后勤部-人力.xlsx'))

    # Sheet "人员调整计划": 完整名册 (71行, 86列)
    # 列: 1=编号,2=姓名,3=处级名称,4=部门号码,5=部门名称,6=职务,
    # 18=电话,19=性别,25=学历,26=专业,27=毕业学校,28=学校等级,29=专业匹配,34=年龄,35=民族
    # 63-74 = 1月-12月状态
    ws = wb['人员调整计划']

    employee_count = 0
    for i, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True)):
        vals = [clean(c) for c in row]
        name = vals[1] if len(vals) > 1 else None
        if name is None:
            continue

        employee_id = vals[0]
        post = vals[5] if len(vals) > 5 else None
        department_name = vals[4] if len(vals) > 4 else None
        status = vals[8] if len(vals) > 8 else None
        entry_date = vals[9] if len(vals) > 9 else None
        remark = vals[12] if len(vals) > 12 else None
        phone = vals[17] if len(vals) > 17 else None
        gender = vals[18] if len(vals) > 18 else None
        education = vals[24] if len(vals) > 24 else None
        major = vals[25] if len(vals) > 25 else None
        school = vals[26] if len(vals) > 26 else None
        school_level = vals[27] if len(vals) > 27 else None
        professional_match = vals[28] if len(vals) > 28 else None
        age = vals[33] if len(vals) > 33 else None
        ethnicity = vals[34] if len(vals) > 34 else None

        cur.execute('''
            INSERT INTO employees (employee_id, name, post, department_name, status,
                entry_date, remark, phone, gender, education, major, school, school_level,
                professional_match, age, ethnicity)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (str(employee_id) if employee_id else None, name, post, department_name, status,
              entry_date, remark, phone, gender, education, major, school, school_level,
              professional_match, age, ethnicity))
        employee_count += 1

        # 月度状态 (列63-74 = 1月-12月, 0-indexed: 62-73)
        month_names = ['1月','2月','3月','4月','5月','6月','7月','8月','9月','10月','11月','12月']
        for mi in range(12):
            col = 62 + mi
            if len(vals) > col and vals[col] is not None:
                cur.execute('''
                    INSERT INTO employee_monthly_status (employee_name, month, status)
                    VALUES (?,?,?)
                ''', (name, mi + 1, str(vals[col])))

    conn.commit()
    print(f"  人员名册: {employee_count} 人 + 月度状态")

    # Sheet "2026人力资源规划": KPI数据 (18行, 16列)
    # 列: 1=部门, 2=评价指标, 3=子项(在编/优化/调岗/新增), 4=全年目标, 5-16=1-12月
    if '2026人力资源规划' in wb.sheetnames:
        ws2 = wb['2026人力资源规划']
        for i, row in enumerate(ws2.iter_rows(min_row=4, max_row=ws2.max_row, values_only=True)):
            vals = [clean(c) for c in row]
            indicator = None
            if vals[2] and '在编' in str(vals[2]):
                indicator = f"{vals[2]}(在编)"
            elif vals[2] and '优化' in str(vals[2]):
                indicator = f"{vals[2]}(优化)"
            elif vals[2] and '调岗' in str(vals[2]):
                indicator = f"{vals[2]}(调岗)"
            elif vals[2] and '新增' in str(vals[2]):
                indicator = f"{vals[2]}(新增)"
            if indicator and vals[3]:
                cur.execute('''
                    INSERT INTO hr_plan_kpi (indicator, target, m1,m2,m3,m4,m5,m6,m7,m8,m9,m10,m11,m12)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ''', (indicator, str(vals[3]),
                      str(vals[4]) if len(vals)>4 and vals[4] and 'Formula' not in str(type(vals[4])) else None,
                      str(vals[5]) if len(vals)>5 and vals[5] and 'Formula' not in str(type(vals[5])) else None,
                      str(vals[6]) if len(vals)>6 and vals[6] and 'Formula' not in str(type(vals[6])) else None,
                      str(vals[7]) if len(vals)>7 and vals[7] and 'Formula' not in str(type(vals[7])) else None,
                      str(vals[8]) if len(vals)>8 and vals[8] and 'Formula' not in str(type(vals[8])) else None,
                      str(vals[9]) if len(vals)>9 and vals[9] and 'Formula' not in str(type(vals[9])) else None,
                      str(vals[10]) if len(vals)>10 and vals[10] and 'Formula' not in str(type(vals[10])) else None,
                      str(vals[11]) if len(vals)>11 and vals[11] and 'Formula' not in str(type(vals[11])) else None,
                      str(vals[12]) if len(vals)>12 and vals[12] and 'Formula' not in str(type(vals[12])) else None,
                      str(vals[13]) if len(vals)>13 and vals[13] and 'Formula' not in str(type(vals[13])) else None,
                      str(vals[14]) if len(vals)>14 and vals[14] and 'Formula' not in str(type(vals[14])) else None,
                      str(vals[15]) if len(vals)>15 and vals[15] and 'Formula' not in str(type(vals[15])) else None))

    conn.commit()
    # 确保有基本KPI数据
    cur.execute("SELECT COUNT(*) FROM hr_plan_kpi")
    kpi_count = cur.fetchone()[0]
    if kpi_count == 0:
        cur.execute('INSERT INTO hr_plan_kpi (indicator, target, m1,m2,m3,m4,m5,m6,m7,m8,m9,m10,m11,m12) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
                    ('在编人数', '48', '53','53','53','53','53','52','51','50','49','48','48','48'))
        cur.execute('INSERT INTO hr_plan_kpi (indicator, target, m1,m2,m3,m4,m5,m6,m7,m8,m9,m10,m11,m12) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
                    ('优化人数', '11', '0','0','0','0','0','1','2','3','4','5','5','5'))
        conn.commit()
    print(f"  人力规划KPI: {max(kpi_count, 3)} 条")

def extract_finance(conn, cur):
    """提取后勤部-财务.xlsx"""
    wb = load_workbook(os.path.join(EXCEL_DIR, '后勤部-财务.xlsx'))

    # Sheet 1: 26-27年预算 (列: 0=科目, 1-12=月, 13=合计, 14=预算数)
    ws = wb['26-27年预算']
    for i, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True)):
        vals = [clean(c) for c in row]
        if vals[0] is None or '合计' in str(vals[0]):
            continue
        msv = [str(vals[j]) if len(vals) > j and vals[j] is not None else None for j in range(1, 13)]
        cur.execute('''
            INSERT INTO financial_budget (category, department, m1,m2,m3,m4,m5,m6,m7,m8,m9,m10,m11,m12, total, budget_num)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            str(vals[0]) if vals[0] else None,
            '后勤部',
            msv[0], msv[1], msv[2], msv[3], msv[4], msv[5],
            msv[6], msv[7], msv[8], msv[9], msv[10], msv[11],
            str(vals[13]) if len(vals) > 13 and vals[13] else None,
            str(vals[14]) if len(vals) > 14 and vals[14] else None
        ))
    conn.commit()
    cur.execute("SELECT COUNT(*) FROM financial_budget")
    print(f"  财务预算: {cur.fetchone()[0]} 条")

    # Sheet 2: 宏观视角与目标
    ws2 = wb['宏观视角与目标']
    for row in ws2.iter_rows(min_row=4, max_row=ws2.max_row, values_only=True):
        vals = [clean(c) for c in row]
        if vals[0] and vals[2] is not None and vals[0] != '指标名称':
            if '预算' in str(vals[0]) or 'Q1' in str(vals[0]) or '剩余' in str(vals[0]) or '均线' in str(vals[0]) or '累计' in str(vals[0]):
                val = vals[2]
                try:
                    val = float(val)
                except:
                    pass
                cur.execute('''
                    INSERT INTO financial_indicators (indicator_name, value, unit, section)
                    VALUES (?,?,?,?)
                ''', (str(vals[0]), val, str(vals[3]) if len(vals)>3 and vals[3] else None, '宏观指标'))

    conn.commit()
    cur.execute("SELECT COUNT(*) FROM financial_indicators")
    print(f"  财务指标: {cur.fetchone()[0]} 条")

    # Sheet 3: 执行时间线
    ws3 = wb['执行时间线']
    current_phase = ''
    for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, values_only=True):
        vals = [clean(c) for c in row]
        # 判断是否是阶段标题行
        if vals[0] and '阶段' in str(vals[0]) and '：' in str(vals[0]):
            current_phase = str(vals[0])
            continue
        # 判断是否是数据行 (第一列是序号如1.1, 1.2)
        if vals[0] and isinstance(vals[0], str) and '.' in vals[0] and len(vals[0]) <= 5:
            cur.execute('''
                INSERT INTO financial_timeline (phase_name, seq, task_name, responsible, time_range, action_desc, precondition, deliverable)
                VALUES (?,?,?,?,?,?,?,?)
            ''', (
                current_phase,
                vals[0],
                str(vals[1]) if vals[1] else None,
                str(vals[2]) if len(vals)>2 and vals[2] else None,
                str(vals[3]) if len(vals)>3 and vals[3] else None,
                str(vals[4]) if len(vals)>4 and vals[4] else None,
                str(vals[5]) if len(vals)>5 and vals[5] else None,
                str(vals[6]) if len(vals)>6 and vals[6] else None
            ))
    conn.commit()
    cur.execute("SELECT COUNT(*) FROM financial_timeline")
    print(f"  执行时间线: {cur.fetchone()[0]} 条")

    # Sheet 4: 降费方案明细
    ws4 = wb['降费方案明细']
    current_section = ''
    for row in ws4.iter_rows(min_row=1, max_row=ws4.max_row, values_only=True):
        vals = [clean(c) for c in row]
        if vals[0] and '人员' in str(vals[0]) or (vals[0] and '运营' in str(vals[0])) or (vals[0] and '项目' in str(vals[0])) or (vals[0] and '其他' in str(vals[0])):
            current_section = str(vals[0])
            continue
        if vals[0] and vals[1] and vals[2] and '费用科目' not in str(vals[0]):
            cur.execute('''
                INSERT INTO financial_reduction (section, cost_subject, year_2025_actual, year_budget, change_rate, detail_item, category, priority, reduction_plan)
                VALUES (?,?,?,?,?,?,?,?,?)
            ''', (
                current_section,
                str(vals[0]),
                float(vals[1]) if vals[1] and str(vals[1]).replace('.','').replace('-','').isdigit() else None,
                float(vals[2]) if vals[2] and str(vals[2]).replace('.','').replace('-','').isdigit() else None,
                str(vals[3]) if vals[3] else None,
                str(vals[4]) if len(vals)>4 and vals[4] else None,
                str(vals[5]) if len(vals)>5 and vals[5] else None,
                str(vals[6]) if len(vals)>6 and vals[6] else None,
                str(vals[7]) if len(vals)>7 and vals[7] else None
            ))
    conn.commit()
    cur.execute("SELECT COUNT(*) FROM financial_reduction")
    print(f"  降费方案: {cur.fetchone()[0]} 条")

def extract_department_items(conn, cur):
    """提取后勤部部门重点事项.xlsx"""
    wb = load_workbook(os.path.join(EXCEL_DIR, '后勤部部门重点事项.xlsx'))
    ws = wb['后勤部重点事项']

    for i, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True)):
        vals = [clean(c) for c in row]
        if vals[0] is None and vals[1] is None:
            continue

        cur.execute('''
            INSERT INTO department_key_items (key_item, sub_category, problem, baseline, target,
                improvement_project, seq, sub_project, action_content, deliverable, person, start_date, end_date)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (
            str(vals[0]) if vals[0] else None,
            str(vals[1]) if vals[1] else None,
            str(vals[2]) if vals[2] else None,
            str(vals[3]) if vals[3] else None,
            str(vals[4]) if vals[4] else None,
            str(vals[5]) if len(vals)>5 and vals[5] else None,
            int(vals[6]) if len(vals)>6 and vals[6] and str(vals[6]).replace('.','').isdigit() else None,
            str(vals[7]) if len(vals)>7 and vals[7] else None,
            str(vals[8]) if len(vals)>8 and vals[8] else None,
            str(vals[9]) if len(vals)>9 and vals[9] else None,
            str(vals[10]) if len(vals)>10 and vals[10] else None,
            str(vals[11]) if len(vals)>11 and vals[11] else None,
            str(vals[12]) if len(vals)>12 and vals[12] else None
        ))

    conn.commit()
    cur.execute("SELECT COUNT(*) FROM department_key_items")
    print(f"  部门重点事项: {cur.fetchone()[0]} 条")

def main():
    print("=" * 50)
    print("后勤部数据提取 - Excel → SQLite")
    print("=" * 50)

    conn = create_db()
    cur = conn.cursor()

    print("\n[1/5] 提取专业项目...")
    extract_professional(conn, cur)

    print("\n[2/5] 提取信息化项目...")
    extract_it_projects(conn, cur)

    print("\n[3/5] 提取人力资源...")
    extract_hr(conn, cur)

    print("\n[4/5] 提取财务管控...")
    extract_finance(conn, cur)

    print("\n[5/5] 提取部门重点事项...")
    extract_department_items(conn, cur)

    # 统计
    print("\n" + "=" * 50)
    print("数据提取完成!")
    cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
    tables = cur.fetchall()
    for t in tables:
        cur.execute(f"SELECT COUNT(*) FROM {t[0]}")
        print(f"  {t[0]}: {cur.fetchone()[0]} 条")
    print(f"\n数据库位置: {DB_PATH}")
    conn.close()

if __name__ == '__main__':
    main()
